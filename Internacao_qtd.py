from datetime import datetime
import APIs.Producao.Api_Censo as api_censo
import pandas as pd
from openpyxl import Workbook, load_workbook


def preencherPlanilha_norte():
    planilha  = r"\\10.0.0.239\automacao_faturamento\CTI\Planilhas\Norte\PLANILHA DE OCUPAÇÃO -2023.xlsx"
    data      = datetime.now()
    dia       = data.day
    mes       = data.month
    linha     = 0
    coluna    = dia
   
    # Encontrar o Mês
    if  (mes == 1):
        linha = linha + 20
    elif(mes == 2):
        linha = linha + 21
    elif(mes == 3):
        linha = linha + 22
    elif(mes == 4):
        linha = linha + 23
    elif(mes == 5):
        linha = linha + 24
    elif(mes == 6):
        linha = linha + 25
    elif(mes == 7):
        linha = linha + 26
    elif(mes == 8):
        linha = linha + 27
    elif(mes == 9):
        linha = linha + 28
    elif(mes == 10):
        linha = linha + 29
    elif(mes == 11):
        linha = linha + 30
    elif(mes == 12):
        linha = linha + 31

    api_censo.auth()

    internacao = api_censo.getInternacao_cti_norte()
    qtd_internado = [internacao]
    df = pd.DataFrame(qtd_internado)
    print(df)
    book = load_workbook(planilha)
    writer = pd.ExcelWriter(planilha, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df.to_excel(writer, 'OCUPAÇÃO UTI', startrow = linha , startcol= coluna, header=False, index=False)
    writer.save()
    print(df)





def preencherPlanilha_sul():
    planilha  = r"\\10.0.0.239\automacao_faturamento\CTI\Planilhas\Sul\ADM X ALTA- OCUPAÇÃO 2023.xlsx"
    data      = datetime.now()
    dia       = data.day
    mes       = data.month
    linha     = 0
    coluna    = dia
   
    # Encontrar o Mês
    if  (mes == 1):
        linha = linha + 20
    elif(mes == 2):
        linha = linha + 21
    elif(mes == 3):
        linha = linha + 22
    elif(mes == 4):
        linha = linha + 23
    elif(mes == 5):
        linha = linha + 24
    elif(mes == 6):
        linha = linha + 25
    elif(mes == 7):
        linha = linha + 26
    elif(mes == 8):
        linha = linha + 27
    elif(mes == 9):
        linha = linha + 28
    elif(mes == 10):
        linha = linha + 29
    elif(mes == 11):
        linha = linha + 30
    elif(mes == 12):
        linha = linha + 31

    api_censo.auth()

    internacao = api_censo.getInternacao_cti_sul()

    qtd_internado = [internacao]
    df = pd.DataFrame(qtd_internado)
    print(df)
    book = load_workbook(planilha)
    writer = pd.ExcelWriter(planilha, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df.to_excel(writer, 'OCUPAÇÃO UTI', startrow = linha , startcol= coluna, header=False, index=False)
    writer.save()


preencherPlanilha_norte()

preencherPlanilha_sul()