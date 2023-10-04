import json
import pandas as pd
from datetime import datetime
import APIs.Producao.Api_Censo as api_censo
from openpyxl import Workbook, load_workbook


def preencherPlanilha_norte():
    # planilha  = r"\\10.0.0.239\automacao_faturamento\CTI\Planilhas\Norte\PLANILHA DE OCUPAÇÃO -2023.xlsx"
    planilha   = r"C:\Users\lucas.timoteo\Desktop\teste cti\ROTINA CTI - 2023.xlsx"
    data       = datetime.now()
    data_atual = data.strftime("%d-%m-%Y")
    dia        = data.day
    mes        = data.month
    linha      = 0
    coluna     = dia
   
    # Encontrar o Mês
    match mes:
        case 1:
            linha = linha + 20
        case 2:
            linha = linha + 21
        case 3:
            linha = linha + 22
        case 4:
            linha = linha + 23
        case 5:
            linha = linha + 24
        case 6:
            linha = linha + 25
        case 7:
            linha = linha + 26
        case 8:
            linha = linha + 27
        case 9:
            linha = linha + 28
        case 10:
            linha = linha + 29
        case 11:
            linha = linha + 30
        case 12:
            linha = linha + 31

    api_censo.auth()

    internacao = api_censo.getInternacao_cti_norte()

    total,pacientes = internacao

    print(total)
    print(pacientes)

    qtd_internado = [total]
    df_final = pd.DataFrame()
    lista = []
    df = pd.read_excel(planilha)
    for i in range(total):

        # Dados obtidos da API do Censo
        date            = [data_atual]
        atendimento     = pacientes[i]['nroRegistroAtendimento']
        nomePaciente    = pacientes[i]['paciente']
        convenio        = pacientes[i]['apelido']
        acomodacao      = pacientes[i]['acomodacao']
        dataPaciente    = pacientes[i]['dataInicial']
        dataAtendimento = str(dataPaciente)
        dataAtendimento = dataAtendimento.replace('T00:00:00','')
        dataAtendimento = dataAtendimento.replace('-','/')
        print(dataAtendimento,type(date))
        dataAtendimento = datetime.strptime(dataAtendimento, '%y/%m/%d').date()

        print(df['ATEND.'])
        atendimentoPlanilha = df['ATEND.'].values.tolist()

        if (atendimento in atendimentoPlanilha):
            continue

        else:
            lista_atualizada = [date, atendimento, nomePaciente, convenio,acomodacao,dataPaciente]
            lista.append(lista_atualizada)

    print(lista)
    dado = pd.DataFrame(lista)
    book = load_workbook(planilha)
    writer = pd.ExcelWriter(planilha, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    # df_internado.to_excel(writer, 'OCUPAÇÃO UTI', startrow = 62 , startcol= 1, header=False, index=False)
    dado.to_excel(writer, 'ADMISSÃO ROTINA', startrow = 62 , startcol= 0, header=False, index=False)
    writer.save()
            





def preencherPlanilha_sul():
    planilha  = r"\\10.0.0.239\automacao_faturamento\CTI\Planilhas\Sul\ADM X ALTA- OCUPAÇÃO 2023.xlsx"
    data      = datetime.now()
    dia       = data.day
    mes       = data.month
    linha     = 0
    coluna    = dia
   
    # Encontrar o Mês
    match mes:
        case 1:
            linha = linha + 20
        case 2:
            linha = linha + 21
        case 3:
            linha = linha + 22
        case 4:
            linha = linha + 23
        case 5:
            linha = linha + 24
        case 6:
            linha = linha + 25
        case 7:
            linha = linha + 26
        case 8:
            linha = linha + 27
        case 9:
            linha = linha + 28
        case 10:
            linha = linha + 29
        case 11:
            linha = linha + 30
        case 12:
            linha = linha + 31

    api_censo.auth()

    internacao = api_censo.getInternacao_cti_sul()

    qtd_internado = [internacao]
    df = pd.DataFrame(qtd_internado)
    print(df)
    book = load_workbook(planilha)
    writer = pd.ExcelWriter(planilha, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df.to_excel(writer, 'OCUPAÇÃO UTI', startrow = linha , startcol= coluna, header=False, index=False)
    writer.save()


preencherPlanilha_norte()

# preencherPlanilha_sul()